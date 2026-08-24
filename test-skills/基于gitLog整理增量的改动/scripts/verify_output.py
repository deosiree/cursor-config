#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验收 commits_raw 与 xlsx：条数、问题根粒度、跨仓抽查。"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    raise SystemExit("请先安装: pip install openpyxl")


def skill_root() -> Path:
    return Path(__file__).resolve().parent.parent


def load_cfg(config_path: Path, meta_root: str | None) -> dict:
    cfg = json.loads(config_path.read_text(encoding="utf-8"))
    root = skill_root()
    meta = Path(meta_root or cfg["metaRoot"])
    if not meta.is_absolute():
        meta = (root / meta).resolve()
    out = meta / cfg["outDir"]
    return {
        "raw": out / cfg.get("rawJsonName", "commits_raw.json"),
        "xlsx": out / cfg["xlsxName"],
        "expectCommits": cfg.get("expectCommits"),
        "expectProblems": cfg.get("expectProblems"),
        "expectSubs": cfg.get("expectSubs"),
    }


def main() -> int:
    p = argparse.ArgumentParser(description="gitLog Excel 验收")
    p.add_argument("--config", default=str(skill_root() / "configs/nebula-huiyan-0707-0807.config.json"))
    p.add_argument("--meta-root")
    p.add_argument("--xlsx", help="覆盖 xlsx 路径（few-shot 验收用）")
    p.add_argument("--raw", help="覆盖 raw json 路径")
    args = p.parse_args()

    cfg_path = Path(args.config)
    if not cfg_path.is_absolute():
        cfg_path = (skill_root() / cfg_path).resolve()
    paths = load_cfg(cfg_path, args.meta_root)
    raw_path = Path(args.raw) if args.raw else paths["raw"]
    xlsx_path = Path(args.xlsx) if args.xlsx else paths["xlsx"]

    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["节点表"]

    types: dict[str, int] = {}
    problem_titles: list[str] = []
    commit_shorts: set[str] = set()
    domain_like = {"租户管理", "菜单管理", "用户管理", "角色管理", "登录鉴权", "权限鉴权", "路由鉴权"}

    for row in ws.iter_rows(min_row=2, values_only=True):
        t = row[2]
        types[t] = types.get(t, 0) + 1
        if t == "问题":
            problem_titles.append(str(row[3] or ""))
        if t == "提交":
            commit_shorts.add(row[6])

    raw_shorts = {c["commit_short"] for c in raw}
    bad_roots = [t for t in problem_titles if t in domain_like]
    report = {
        "rawCount": len(raw),
        "commitRows": len(commit_shorts),
        "problemRoots": types.get("问题", 0),
        "subProblems": types.get("子问题", 0),
        "missingInExcel": sorted(raw_shorts - commit_shorts),
        "extraInExcel": sorted(commit_shorts - raw_shorts),
        "domainLikeProblemRoots": bad_roots,
        "passed": (
            len(raw_shorts - commit_shorts) == 0
            and len(commit_shorts - raw_shorts) == 0
            and not bad_roots
        ),
    }
    full_cfg = json.loads(cfg_path.read_text(encoding="utf-8"))
    for key, field in [
        ("expectCommits", "rawCount"),
        ("expectProblems", "problemRoots"),
        ("expectSubs", "subProblems"),
    ]:
        if full_cfg.get(key) is not None and report[field] != full_cfg[key]:
            report["passed"] = False
            report.setdefault("expectMismatch", []).append(f"{field}: got {report[field]}, want {full_cfg[key]}")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    sys.exit(main())
