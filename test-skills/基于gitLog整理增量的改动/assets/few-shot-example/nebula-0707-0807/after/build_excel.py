#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""域名标注 + 主题问题树 + Excel 导出（--config 驱动）。"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("请先安装: pip install openpyxl")


NODE_COLUMNS = [
    "节点ID", "父节点ID", "节点类型", "标题", "仓库",
    "提交完整ID", "提交短ID", "提交标题", "提交完整信息", "提交日期",
    "是否主域", "域名标签", "域名映射依据", "域名主责人", "我的角色",
    "差分同题说明", "改动文件摘要", "同级排序", "内测用例提示",
]
DICT_COLUMNS = ["域名标签", "是否惠岩主责主域", "域名主责人", "典型仓库", "映射线索", "备注"]


def skill_root() -> Path:
    return Path(__file__).resolve().parent.parent


def resolve_cfg_path(base: Path, rel: str) -> Path:
    p = Path(rel)
    return p if p.is_absolute() else (base / p).resolve()


def load_bundle(config_path: Path, meta_root_override: str | None) -> dict[str, Any]:
    root = skill_root()
    cfg = json.loads(config_path.read_text(encoding="utf-8"))
    if meta_root_override:
        cfg["metaRoot"] = meta_root_override
    meta = Path(cfg["metaRoot"])
    if not meta.is_absolute():
        meta = (root / meta).resolve()
    cfg["_metaRoot"] = meta
    cfg["_outDir"] = (meta / cfg["outDir"]).resolve()
    cfg["_repos"] = {k: (meta / v).resolve() for k, v in cfg["repos"].items()}
    cfg["_domainDict"] = json.loads(
        resolve_cfg_path(root, cfg["domainDictFile"]).read_text(encoding="utf-8")
    )
    theme = json.loads(resolve_cfg_path(root, cfg["themeRulesFile"]).read_text(encoding="utf-8"))
    cfg["_themeTitle"] = theme["themeTitle"]
    cfg["_clusterRules"] = [(r[0], r[1]) for r in theme["clusterRules"]]
    cfg["_themeGroups"] = json.loads(
        resolve_cfg_path(root, cfg["themeGroupsFile"]).read_text(encoding="utf-8")
    )
    cfg["_themeToGroup"] = {
        t: gid for gid, meta in cfg["_themeGroups"].items() for t in meta["subs"]
    }
    return cfg


def parse_scope(subject: str) -> tuple[str, str]:
    m = re.match(r"^(feat|fix|refactor|chore|revert|docs|style|test|perf)\(([^)]+)\):\s*(.+)$", subject)
    if m:
        return m.group(2), m.group(3)
    m2 = re.match(r"^(feat|fix|refactor|chore|revert|docs|style|test|perf):\s*(.+)$", subject)
    if m2:
        return "", m2.group(2)
    return "", subject


def tag_domain(c: dict, cfg: dict[str, Any]) -> dict[str, str]:
    """按 config 默认主责人与协作域标注。"""
    owner = cfg.get("defaultOwner", cfg.get("author", "惠岩"))
    collab = cfg.get("collaborators") or {}
    subject = c["subject"]
    files = " ".join(c.get("files") or [])
    scope, _ = parse_scope(subject)
    text = f"{subject} {files}".lower()
    repo = c["repo"]

    def row(label: str, primary: str, basis: str, dom_owner: str, role: str) -> dict[str, str]:
        return {
            "是否主域": primary,
            "域名标签": label,
            "域名映射依据": basis,
            "域名主责人": dom_owner,
            "我的角色": role,
        }

    if repo == "nebula-ui" or "nesecret" in text or "guardedsecret" in text:
        return row("密码框", "是", f"scope={scope}; 密码框", owner, "主责")
    if scope == "i18n" or "nei18n" in text or "国际化" in subject:
        cinfo = collab.get("国际化", {})
        return row("国际化", "否", f"scope={scope}; i18n", cinfo.get("owner", "叶倩"), cinfo.get("myRole", "协作接入"))
    if "路由鉴权" in subject or "路由守卫" in subject or (
        scope == "perm" and ("pageurl" in text or "funcurl" in text)
    ) or (scope == "auth" and "多子路由" in subject):
        cinfo = collab.get("路由鉴权", {})
        role = "辅助" if "多子路由" in subject else cinfo.get("myRole", "协作")
        return row("路由鉴权", "否", f"scope={scope}; 路由", cinfo.get("owner", "杨欣静"), role)
    if scope in ("perm", "iam") or "查写二分" in subject or "v-hasperm" in text or "隐藏的权限" in subject:
        return row("权限鉴权", "是", f"scope={scope}; 权限", owner, "主责")
    if "401" in subject or "会话过期" in subject:
        label = "登录鉴权" if repo == "microfb" else "通用视图/表单"
        return row(label, "是" if scope in ("auth", "request") else "否", "401", owner, "主责")

    scope_map = {
        "auth": "登录鉴权", "login": "登录鉴权", "tenant": "租户管理", "user": "用户管理",
        "role": "角色管理", "menu": "菜单管理", "profile": "个人中心",
        "securityConfig": "安全配置", "securityconfig": "安全配置",
        "seccenter": "租户管理", "resource": "租户管理", "api": "API通道",
    }
    primary_labels = {"登录鉴权", "租户管理", "用户管理", "角色管理", "菜单管理", "个人中心", "安全配置"}
    if scope in scope_map:
        label = scope_map[scope]
        return row(label, "是" if label in primary_labels else "否", f"scope={scope}", owner, "主责" if label in primary_labels else "协作")

    hints = [
        ("views/tenant", "租户管理", "租户"), ("views/system/role", "角色管理", "角色"),
        ("views/system/menu", "菜单管理", "菜单"), ("views/system/user", "用户管理", "用户"),
        ("views/profile", "个人中心", "profile"), ("views/login", "登录鉴权", "登录"),
    ]
    for path_hint, label, kw in hints:
        if path_hint in files or kw in subject:
            return row(label, "是", path_hint, owner, "主责")
    if scope in ("types",) or "vue-tsc" in subject or scope == "build" or "内测临时" in subject:
        return row("工程化/类型", "否", "types/build", owner, "主责")
    if "direct" in subject or "forward" in subject:
        return row("API通道", "否", "direct/forward", owner, "主责")
    return row("通用视图/表单", "否", f"scope={scope or '无'}", owner, "主责")


def cluster_key(c: dict, cfg: dict[str, Any]) -> str:
    subject = c["subject"]
    for key, pat in cfg["_clusterRules"]:
        if re.search(pat, subject, re.I):
            return key
    s, _ = parse_scope(subject)
    return f"未归类-{s or 'misc'}"


def git_diff_note(c: dict, prev: dict | None, repos: dict[str, Path]) -> str:
    repo_path = repos.get(c["repo"])
    summary = c.get("stat_summary", "")
    if repo_path:
        try:
            stat = subprocess.check_output(
                ["git", "show", "--stat", "--pretty=format:", c["commit_id"]],
                cwd=repo_path, encoding="utf-8", errors="replace",
            )
            lines = [l.strip() for l in stat.splitlines() if l.strip()]
            summary = lines[-1] if lines else summary
        except Exception:
            pass
    _, desc = parse_scope(c["subject"])
    note = f"本版：{desc[:120]}"
    if summary:
        note += f"；变更：{summary[:150]}"
    if prev and prev.get("subject"):
        note += f"；相对上一版({prev['commit_short']})为同主题迭代"
    return note


def commit_node(c: dict, tag: dict, parent: str, sort: int, prev: dict | None, repos: dict) -> dict:
    files = c.get("files") or []
    fs = "; ".join(files[:8])
    if len(files) > 8:
        fs += f" …共{len(files)}个文件"
    _, desc = parse_scope(c["subject"])
    return {
        "节点ID": f"C-{c['commit_short']}",
        "父节点ID": parent,
        "节点类型": "提交",
        "标题": desc[:200] if desc else c["subject"],
        "仓库": c["repo"],
        "提交完整ID": c["commit_id"],
        "提交短ID": c["commit_short"],
        "提交标题": c["subject"],
        "提交完整信息": c["message"],
        "提交日期": c["date"],
        "是否主域": tag["是否主域"],
        "域名标签": tag["域名标签"],
        "域名映射依据": tag["域名映射依据"],
        "域名主责人": tag["域名主责人"],
        "我的角色": tag["我的角色"],
        "差分同题说明": git_diff_note(c, prev, repos),
        "改动文件摘要": fs,
        "同级排序": sort,
        "内测用例提示": tag["域名标签"],
    }


def build_problem_tree(commits: list[dict], tagged: dict[str, dict], cfg: dict[str, Any]) -> list[dict]:
    clusters: dict[str, list[dict]] = defaultdict(list)
    for c in commits:
        clusters[cluster_key(c, cfg)].append(c)

    units: dict[str, dict[str, Any]] = {}
    for theme, group in clusters.items():
        gid = cfg["_themeToGroup"].get(theme)
        if gid:
            meta = cfg["_themeGroups"][gid]
            if gid not in units:
                units[gid] = {"title": meta["title"], "themes": [], "grouped": True, "sub_titles": meta["subs"]}
            units[gid]["themes"].append(theme)
        else:
            title = cfg["_themeTitle"].get(theme, theme)
            if theme.startswith("未归类-"):
                title = f"未归类（{theme[4:]}）"
            units[theme] = {"title": title, "themes": [theme], "grouped": False, "sub_titles": {}}

    def sort_key(item: tuple[str, dict]) -> tuple:
        gid, u = item
        dates = [c["date"] for t in u["themes"] for c in clusters[t]]
        return (min(dates) if dates else "9999", gid)

    nodes: list[dict] = []
    for prob_idx, (gid, unit) in enumerate(sorted(units.items(), key=sort_key), 1):
        pid = f"P{prob_idx:03d}"
        themes = sorted(unit["themes"])
        all_c: list[dict] = []
        for t in themes:
            all_c.extend(clusters[t])
        all_c.sort(key=lambda x: (x["date"], x["repo"], x["commit_short"]))
        if not all_c:
            continue
        tag0 = tagged[all_c[0]["commit_short"]]
        use_subs = unit["grouped"] and len(themes) >= 2
        nodes.append({
            "节点ID": pid, "父节点ID": "", "节点类型": "问题", "标题": unit["title"],
            "仓库": "+".join(sorted({c["repo"] for c in all_c})),
            "是否主域": tag0["是否主域"], "域名标签": tag0["域名标签"],
            "域名映射依据": f"主题={'+'.join(themes)}",
            "域名主责人": tag0["域名主责人"], "我的角色": tag0["我的角色"],
            "提交日期": all_c[0]["date"], "同级排序": prob_idx,
        })
        if use_subs:
            for si, theme in enumerate(themes, 1):
                group = sorted(clusters[theme], key=lambda x: (x["date"], x["commit_short"]))
                sid = f"{pid}-S{si:02d}"
                tag = tagged[group[0]["commit_short"]]
                nodes.append({
                    "节点ID": sid, "父节点ID": pid, "节点类型": "子问题",
                    "标题": unit["sub_titles"].get(theme, cfg["_themeTitle"].get(theme, theme)),
                    "仓库": "+".join(sorted({c["repo"] for c in group})),
                    "是否主域": tag["是否主域"], "域名标签": tag["域名标签"],
                    "域名映射依据": f"子主题={theme}",
                    "域名主责人": tag["域名主责人"], "我的角色": tag["我的角色"],
                    "提交日期": group[0]["date"], "同级排序": si,
                })
                prev = None
                for ci, c in enumerate(group, 1):
                    nodes.append(commit_node(c, tagged[c["commit_short"]], sid, ci, prev, cfg["_repos"]))
                    prev = c
        else:
            if unit["grouped"] and len(themes) == 1:
                only = themes[0]
                nodes[-1]["标题"] = unit["sub_titles"].get(only, unit["title"])
            prev = None
            for ci, c in enumerate(all_c, 1):
                nodes.append(commit_node(c, tagged[c["commit_short"]], pid, ci, prev, cfg["_repos"]))
                prev = c
    return nodes


def write_excel(nodes: list[dict], domain_dict: list[dict], out_xlsx: Path, cfg: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "节点表"
    hf = PatternFill("solid", fgColor="4472C4")
    hfont = Font(color="FFFFFF", bold=True)
    for col, name in enumerate(NODE_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill, cell.font = hf, hfont
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for ri, node in enumerate(nodes, 2):
        for ci, name in enumerate(NODE_COLUMNS, 1):
            ws.cell(row=ri, column=ci, value=node.get(name, "")).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
    ws2 = wb.create_sheet("域名字典")
    for col, name in enumerate(DICT_COLUMNS, 1):
        cell = ws2.cell(row=1, column=col, value=name)
        cell.fill, cell.font = hf, hfont
    for ri, row in enumerate(domain_dict, 2):
        for ci, name in enumerate(DICT_COLUMNS, 1):
            ws2.cell(row=ri, column=ci, value=row.get(name, ""))
    ws3 = wb.create_sheet("使用说明")
    n_commit = sum(1 for n in nodes if n.get("节点类型") == "提交")
    n_prob = sum(1 for n in nodes if n.get("节点类型") == "问题")
    n_sub = sum(1 for n in nodes if n.get("节点类型") == "子问题")
    lines = [
        f"{cfg.get('author')} {cfg.get('since')} 起提交整理",
        "问题根=主题，域名只做列；子问题仅路由鉴权/密码框主题组",
        f"共 {n_commit} 提交，{n_prob} 问题根，{n_sub} 子问题",
    ]
    for i, line in enumerate(lines, 1):
        ws3.cell(row=i, column=1, value=line)
    widths = [14, 14, 8, 40, 12, 42, 10, 50, 60, 12, 8, 12, 30, 10, 10, 50, 40, 8, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"wrote {out_xlsx}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default=str(skill_root() / "configs/nebula-huiyan-0707-0807.config.json"))
    parser.add_argument("--meta-root", help="覆盖 metaRoot")
    args = parser.parse_args()
    cfg_path = Path(args.config)
    if not cfg_path.is_absolute():
        cfg_path = (skill_root() / cfg_path).resolve()
    cfg = load_bundle(cfg_path, args.meta_root)
    raw_path = cfg["_outDir"] / cfg.get("rawJsonName", "commits_raw.json")
    commits = json.loads(raw_path.read_text(encoding="utf-8"))
    tagged = {c["commit_short"]: tag_domain(c, cfg) for c in commits}
    nodes = build_problem_tree(commits, tagged, cfg)
    xlsx = cfg["_outDir"] / cfg["xlsxName"]
    write_excel(nodes, cfg["_domainDict"], xlsx, cfg)
    print(f"nodes={len(nodes)} commits={sum(1 for n in nodes if n.get('节点类型')=='提交')}")


if __name__ == "__main__":
    main()
